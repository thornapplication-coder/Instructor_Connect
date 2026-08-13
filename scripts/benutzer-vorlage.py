#!/usr/bin/env python3
"""Excel-Vorlage fuer den Benutzer-Massenimport bauen.

    python3 scripts/benutzer-vorlage.py

Warum ein Generator und keine von Hand gepflegte Datei: Die Spalten stehen
zweimal — in `IMPORT_COLUMNS` (src/userImport.ts, daraus entsteht die
CSV-Vorlage) und in der Excel-Datei. Zwei Prüfungen für dieselbe Sache laufen
auseinander; hier waren es zwei Quellen für dieselben Spalten. Dieses Skript
liest die Spalten und die Musterliste aus dem Quelltext und baut die Datei
daraus, statt sie zu wiederholen. Wer eine Spalte ändert, führt es erneut aus.

Braucht `openpyxl` (nur zum Bauen — die App selbst liest kein Excel, siehe
den Kommentar in src/userImport.ts).
"""
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

WURZEL = Path(__file__).resolve().parent.parent
ZIEL = WURZEL / 'public' / 'Instructor-Connect-Benutzer-Vorlage.xlsx'

TINTE = '1F3864'
GRAU = '7F7F7F'
HELLGRAU = 'EEEEEE'
SCHRIFT = 'Arial'


def _liste_aus_quelle(datei: Path, muster: str) -> list[str]:
    """Ein String-Array aus dem TypeScript-Quelltext ziehen."""
    text = datei.read_text(encoding='utf-8')
    treffer = re.search(muster, text, re.S)
    if not treffer:
        sys.exit(f'FEHLER: {muster!r} nicht gefunden in {datei}')
    return re.findall(r"'([^']+)'", treffer.group(1))


SPALTEN = _liste_aus_quelle(WURZEL / 'src' / 'userImport.ts', r'IMPORT_COLUMNS = \[(.*?)\] as const')
MUSTER = _liste_aus_quelle(WURZEL / 'src' / 'sandbox' / 'seed.ts', r'\n      aircraftTypes: \[([^\]]*)\],')

# Die Musterliste ist im Admin-Panel aenderbar. Die Datei kann deshalb nur den
# Auslieferungsstand zeigen — darauf weist das Blatt „Muster" ausdruecklich hin,
# sonst traegt jemand ein Muster ein, das seine ATO gar nicht fuehrt.
if not MUSTER:
    sys.exit('FEHLER: Musterliste aus seed.ts ist leer')

ROLLEN = ['superadmin', 'admin', 'training admin', 'member']

# Spaltenbeschreibungen. Der Text zur Musterspalte ist der Grund fuer diese
# Fassung: Vorher stand dort „Leer = keines" — seit die Sichtbarkeit am Muster
# haengt, waere ein Mitglied ohne Muster fuer nichts zustaendig und saehe
# weder Lesson Plan noch Info noch Chat.
SPALTEN_TEXT = {
    'Name': 'Vor- und Nachname. Pflichtfeld.',
    'Email': 'Anmeldeadresse. Pflichtfeld, muss eine freigegebene Domain haben.',
    'Phone': 'Optional, beliebiges Format.',
    'Role': f'{" | ".join(ROLLEN)}. Pflichtfeld.',
    'AircraftTypes': (
        'Muster, mehrere mit Semikolon: CL30; C560 XLS+. '
        'PFLICHT fuer member und admin — davon haengt ab, was sie sehen. '
        'Bei superadmin und training admin darf die Zelle leer bleiben: '
        'Diese beiden sehen ohnehin alles. Schreibweise siehe Blatt „Muster".'
    ),
    'CanGrade': 'Darf Grading-Formulare ausfuellen. ja/nein.',
    'IsTrainee': 'Erscheint in der Pilotenauswahl. ja/nein.',
    'CanEditDirectory': 'Darf Who-to-call pflegen. ja/nein.',
    'Active': 'Angemeldet werden kann sich nur mit ja.',
}

ROLLEN_TEXT = {
    'superadmin': 'Vollzugriff inklusive Rechte-Matrix, Einstellungen und endgueltigem Loeschen. Sieht alle Muster.',
    'admin': 'Gruppen-Admin: Formulare, Instructor Info, Lesson Plans, Chats verwalten — fuer die eigenen Muster.',
    'training admin': 'Nur-lesender Zugriff auf die gesamte Formularablage samt Download. Sieht alle Muster.',
    'member': 'Instruktor: eigene Formulare, Chats, Info, Notizen — fuer die eigenen Muster.',
}


def _titel(ws, zelle, text, groesse=14):
    ws[zelle] = text
    ws[zelle].font = Font(name=SCHRIFT, size=groesse, bold=True, color=TINTE)


def _zeile(ws, r, links, rechts):
    ws.cell(r, 2, links).font = Font(name=SCHRIFT, size=10, bold=True)
    z = ws.cell(r, 3, rechts)
    z.font = Font(name=SCHRIFT, size=10)
    z.alignment = Alignment(wrap_text=True, vertical='top')


def blatt_anleitung(wb):
    ws = wb.create_sheet('Anleitung')
    ws.sheet_view.showGridLines = False
    for spalte, breite in (('A', 4), ('B', 26), ('C', 88)):
        ws.column_dimensions[spalte].width = breite

    _titel(ws, 'B2', 'Instructor Connect — Benutzer anlegen')
    ws['B3'] = 'Vorlage fuer den Massenimport im Admin-Panel'
    ws['B3'].font = Font(name=SCHRIFT, size=10, color=GRAU)

    r = 5
    for links, rechts in (
        ('1. Ausfuellen', 'Im Blatt „Benutzer“ ab Zeile 4 eintragen — eine Person je Zeile. Die beiden grauen Beispielzeilen (2 und 3) zeigen das Format; sie duerfen stehen bleiben oder geloescht werden.'),
        ('2. Speichern', 'Datei → Speichern unter → Dateityp „CSV UTF-8 (durch Trennzeichen getrennt) (*.csv)“. Wichtig: UTF-8, sonst werden Umlaute falsch.'),
        ('3. Hochladen', 'In der App: Admin Panel → Benutzer → „Benutzer importieren“ → Datei waehlen.'),
        ('4. Pruefen', 'Vor dem Anlegen zeigt die App jede Zeile mit Status. Angelegt werden nur die fehlerfreien Zeilen — die uebrigen koennen in der Datei korrigiert und erneut hochgeladen werden.'),
    ):
        _zeile(ws, r, links, rechts)
        r += 1

    r += 1
    _titel(ws, f'B{r}', 'Muster (Aircraft Types)', 12)
    r += 1
    for links, rechts in (
        ('Warum das zaehlt', 'Das Muster entscheidet, was eine Person inhaltlich sieht: Lesson Plans, Instructor Info und Chats. Wer keinem Muster zugeordnet ist, sieht davon nichts.'),
        ('Pflicht fuer', 'member und admin. Eine solche Zeile ohne Muster weist der Import ab — sie ergaebe ein Konto, das fuer nichts zustaendig ist.'),
        ('Frei fuer', 'superadmin und training admin. Beide sehen alle Muster, unabhaengig von der Zelle; eine Eintragung aendert daran nichts.'),
        ('Mehrere', 'Mit Semikolon trennen: CL30; C560 XLS+. Kein Schraegstrich — Musterkennungen enthalten selbst welche („ATR 42/72“).'),
        ('Schreibweise', 'Muss der Liste im Blatt „Muster“ entsprechen. Unbekannte Muster halten die Zeile nicht auf, werden aber verworfen und sind danach im Admin-Panel nachzutragen.'),
        ('Nicht betroffen', 'Formularablage, Statistik und Behoerdenexport folgen der Rolle, nicht dem Muster — die Aufbewahrungspflicht gilt fuer den ganzen Bestand.'),
    ):
        _zeile(ws, r, links, rechts)
        r += 1

    r += 1
    _titel(ws, f'B{r}', 'Anmeldung', 12)
    r += 1
    for links, rechts in (
        ('Kein Passwort', 'Angemeldet wird sich mit der E-Mail-Adresse und einem 6-stelligen Code. Es gibt deshalb keine Passwortspalte — die Adresse ist die Anmeldung.'),
        ('Domain', 'Die Domain der Adresse muss im Admin-Panel unter Einstellungen → Erlaubte Domains stehen. Sonst kann sich die Person nie anmelden, und der Import weist die Zeile ab.'),
        ('Aktiv', 'Nur Personen mit Active = ja koennen sich anmelden. So legt man jemanden im Voraus an, ohne ihn schon freizuschalten.'),
    ):
        _zeile(ws, r, links, rechts)
        r += 1

    r += 1
    _titel(ws, f'B{r}', 'Spalten', 12)
    r += 1
    for spalte in SPALTEN:
        _zeile(ws, r, spalte, SPALTEN_TEXT[spalte])
        r += 1

    r += 1
    _titel(ws, f'B{r}', 'Rollen', 12)
    r += 1
    for rolle in ROLLEN:
        _zeile(ws, r, rolle, ROLLEN_TEXT[rolle])
        r += 1
    return ws


def blatt_muster(wb):
    """Die gueltigen Musterkennungen zum Abschreiben — und als Auswahlliste."""
    ws = wb.create_sheet('Muster')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 88

    _titel(ws, 'A1', 'Gueltige Muster')
    hinweis = ws['A2']
    hinweis.value = (
        'Stand der Auslieferung. Die Liste ist im Admin-Panel unter Einstellungen aenderbar — '
        'gilt fuer die eigene ATO die dortige Liste, nicht diese. Schreibweise genau uebernehmen; '
        'Gross- und Kleinschreibung ist dabei egal.'
    )
    hinweis.font = Font(name=SCHRIFT, size=10, color=GRAU)
    hinweis.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A2:B2')
    ws.row_dimensions[2].height = 30

    for i, muster in enumerate(MUSTER, start=4):
        z = ws.cell(i, 1, muster)
        z.font = Font(name=SCHRIFT, size=10)
    return ws


def blatt_benutzer(wb):
    ws = wb.create_sheet('Benutzer')
    ws.freeze_panes = 'A2'
    breiten = {'Name': 26, 'Email': 34, 'Phone': 20, 'Role': 16, 'AircraftTypes': 26}
    for i, spalte in enumerate(SPALTEN, start=1):
        z = ws.cell(1, i, spalte)
        z.font = Font(name=SCHRIFT, size=11, bold=True, color=TINTE)
        ws.column_dimensions[z.column_letter].width = breiten.get(spalte, 12 if len(spalte) < 12 else 18)

    zwei = '; '.join(MUSTER[:2]) if len(MUSTER) >= 2 else (MUSTER[0] if MUSTER else '')
    beispiele = [
        ['Max Beispiel', 'max.beispiel@aviationacademy.at', '+43 664 1234567', 'member', zwei, 'ja', 'nein', 'nein', 'ja'],
        ['Erika Beispiel', 'erika.beispiel@aviationacademy.at', '+43 664 7654321', 'admin', MUSTER[0] if MUSTER else '', 'ja', 'nein', 'ja', 'ja'],
    ]
    for r, werte in enumerate(beispiele, start=2):
        for c, wert in enumerate(werte, start=1):
            z = ws.cell(r, c, wert)
            z.font = Font(name=SCHRIFT, size=10, color=GRAU)
            z.fill = PatternFill('solid', fgColor=HELLGRAU)

    letzte = 203
    dv_rolle = DataValidation(type='list', formula1=f'"{",".join(ROLLEN)}"', allow_blank=True)
    dv_ja = DataValidation(type='list', formula1='"ja,nein"', allow_blank=True)
    ws.add_data_validation(dv_rolle)
    ws.add_data_validation(dv_ja)
    for i, spalte in enumerate(SPALTEN, start=1):
        buchstabe = ws.cell(1, i).column_letter
        if spalte == 'Role':
            dv_rolle.add(f'{buchstabe}2:{buchstabe}{letzte}')
        elif spalte in ('CanGrade', 'IsTrainee', 'CanEditDirectory', 'Active'):
            dv_ja.add(f'{buchstabe}2:{buchstabe}{letzte}')
        elif spalte == 'AircraftTypes':
            # Bewusst KEINE Auswahlliste: Excel laesst je Zelle nur einen
            # Listeneintrag zu, die Zuordnung ist aber eine Mehrfachauswahl.
            # Eine Liste haette hier also genau das verboten, was die Spalte
            # koennen muss. Stattdessen ein Hinweis, der beim Anklicken steht.
            # Ohne `type`: keine Pruefung, nur der Sprechblasen-Hinweis beim
            # Anklicken. Eine `custom`-Regel mit Formel TRUE taete dasselbe,
            # traegt aber eine Formel in die Datei, die nichts pruefen soll.
            dv_muster = DataValidation(
                allow_blank=True,
                showInputMessage=True, showErrorMessage=False,
                promptTitle='Muster', prompt=(
                    'Mehrere mit Semikolon trennen: CL30; C560 XLS+\n'
                    'Pflicht fuer member und admin.\n'
                    'Gueltige Kennungen im Blatt „Muster".'
                ),
            )
            ws.add_data_validation(dv_muster)
            dv_muster.add(f'{buchstabe}2:{buchstabe}{letzte}')
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    # Feste Kopfdaten: openpyxl stempelt sonst die aktuelle Uhrzeit in
    # docProps/core.xml, und jeder Lauf ergaebe eine geaenderte Datei, ohne
    # dass sich am Inhalt etwas geaendert haette. Dieselbe Ueberlegung wie
    # beim reproduzierbaren Build — ein Diff soll etwas bedeuten.
    wb.properties.creator = 'Instructor Connect'
    wb.properties.created = wb.properties.modified = datetime(2026, 1, 1)
    blatt_anleitung(wb)
    blatt_benutzer(wb)
    blatt_muster(wb)
    wb.active = 0
    ZIEL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ZIEL)
    print(f'geschrieben: {ZIEL.relative_to(WURZEL)}')
    print(f'  Spalten aus src/userImport.ts: {", ".join(SPALTEN)}')
    print(f'  Muster aus src/sandbox/seed.ts: {", ".join(MUSTER)}')


if __name__ == '__main__':
    main()
